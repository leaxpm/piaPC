import logging
try:
    from pyhunter import PyHunter
    from openpyxl import Workbook
except:
    logging.error('Necesita las librerias pyhunter y openpyxl \n pip install pyhunter \n pip install openpyxl')

#hunter = PyHunter(input('Ingresa tu api: \n'))

def Busqueda(organizacion):
    """
        **Search**
        This Module find the given Enterprise emails with Hunter API
    """
    # Cantidad de resultados esperados de la búsqueda
    busqueda = 1
    # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(company=organizacion, limit=busqueda) 
    return resultado


def GuardarInformacion(datosEncontrados, organizacion):
    """
        **SavesInfo**
        This Module saves the retrieved info into a xlsx file
    """
    libro = Workbook()
    #Creando sheet en excel
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    hoja.cell(1, 1, "Dominio")
    hoja.cell(1, 2, datosEncontrados['domain'])
    hoja.cell(2, 1, "Organización")
    hoja.cell(2, 2, datosEncontrados['organization'])
    hoja.cell(3, 1, "Correo")
    hoja.cell(3, 2, datosEncontrados['emails'][0]['value'])
    hoja.cell(4, 1, "Nombre(s)")
    hoja.cell(4, 2, datosEncontrados['emails'][0]['first_name'])
    hoja.cell(5, 1, "Apellidos")
    hoja.cell(5, 2, datosEncontrados['emails'][0]['last_name'])
    #Guardando la informacion recaudada
    libro.save("Hunter" + organizacion + ".xlsx")  

#Inicio
def main(apiKey,organizacion):
    #Preguntando Target
    global hunter 
    hunter = PyHunter(apiKey)
    datosEncontrados = Busqueda(organizacion)
    if datosEncontrados:
        exit()
    else:
        #print(datosEncontrados)
        GuardarInformacion(datosEncontrados, organizacion)
        print("Datos Guardados en Hunter"+organizacion+".xlsx")
